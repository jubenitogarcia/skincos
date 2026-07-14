from __future__ import annotations

from pathlib import Path
import re
import unicodedata
from datetime import date, datetime, time

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


DATE_FMT = "dd/mm/yyyy"
TIME_FMT = "hh:mm"
# Quote the currency symbol so spreadsheet apps don't treat "$" as a token.
CURRENCY_FMT = '"R$" #,##0.00'
INTEGER_FMT = "0"
TEXT_FMT = "@"


def _format_brl(value: float) -> str:
    # Format as: R$ 1.234,56 (string only, for width estimation).
    s = f"{value:,.2f}"
    s = s.replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {s}"


def _display_value(value: object, col_type: str | None) -> str:
    if value is None:
        return ""
    if col_type == "date":
        if isinstance(value, datetime):
            return value.strftime("%d/%m/%Y")
        if isinstance(value, date):
            return value.strftime("%d/%m/%Y")
    if col_type == "time":
        if isinstance(value, time):
            return value.strftime("%H:%M")
    if col_type == "currency":
        if isinstance(value, (int, float)):
            return _format_brl(float(value))
    if col_type == "int":
        if isinstance(value, (int, float)):
            return str(int(value))
    return str(value)


def _is_empty_value(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def _norm_header(value: object) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    text = "".join(
        ch
        for ch in unicodedata.normalize("NFD", text)
        if unicodedata.category(ch) != "Mn"
    )
    text = re.sub(r"\s+", " ", text).upper()
    return text


DATE_HEADERS = {"DATA", "DIA", "DATA INICIAL", "DATA FINAL"}
TIME_HEADERS = {"HORARIO", "HORÁRIO"}
CURRENCY_HEADERS = {
    "VALOR",
    "VALOR PAGO",
    "VALOR TOTAL",
    "CREDITO CLIENTE",
    "CRÉDITO CLIENTE",
}
INTEGER_HEADERS = {"QUANTIDADE", "PARCELAS"}
TEXT_HEADERS = {
    "CLIENTE",
    "PAGAMENTO",
    "FORMA DE PAGAMENTO",
    "STATUS",
    "TIPO DE AGENDAMENTO",
    "PROFISSIONAL",
    "TELEFONE",
    "CPF",
    "POR ONDE NOS CONHECEU",
    "SERVICO A REALIZAR",
    "SERVIÇO A REALIZAR",
    "OBSERVACOES",
    "OBSERVAÇÕES",
    "TITULO",
    "TÍTULO",
}


def _parse_float(value: str) -> float | None:
    raw = value.replace("R$", "").replace(".", "").replace(",", ".").strip()
    if not raw:
        return None
    try:
        return float(raw)
    except Exception:
        return None


def _parse_date(value: str) -> date | None:
    raw = value.strip()
    if not raw:
        return None
    try:
        return datetime.strptime(raw, "%d/%m/%Y").date()
    except Exception:
        return None


def _parse_time(value: str) -> time | None:
    raw = value.strip()
    if not raw:
        return None
    for fmt in ("%H:%M", "%H:%M:%S"):
        try:
            return datetime.strptime(raw, fmt).time()
        except Exception:
            continue
    return None


def _is_dash(value: object) -> bool:
    return isinstance(value, str) and value.strip() == "–"


def format_workbook(path: Path) -> None:
    wb = load_workbook(path)
    font = Font(name="Helvetica", size=12)
    for ws in wb.worksheets:
        max_row = ws.max_row or 1
        max_col = ws.max_column or 1
        last_row = 1
        last_col = 1
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                if not _is_empty_value(cell.value):
                    if cell.row > last_row:
                        last_row = cell.row
                    if cell.column > last_col:
                        last_col = cell.column
        if last_row < max_row:
            ws.delete_rows(last_row + 1, max_row - last_row)
        if last_col < max_col:
            ws.delete_cols(last_col + 1, max_col - last_col)

        max_row = ws.max_row or 1
        max_col = ws.max_column or 1
        ws.freeze_panes = "A2"
        header = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
        col_types: dict[int, str] = {}
        for idx, h in enumerate(header, start=1):
            key = _norm_header(h)
            if not key:
                continue
            if key in DATE_HEADERS:
                col_types[idx] = "date"
            elif key in TIME_HEADERS:
                col_types[idx] = "time"
            elif key in CURRENCY_HEADERS:
                col_types[idx] = "currency"
            elif key in INTEGER_HEADERS:
                col_types[idx] = "int"
            elif key in TEXT_HEADERS:
                col_types[idx] = "text"

        max_widths: dict[int, int] = {i: 0 for i in range(1, max_col + 1)}
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.font = font
                if cell.row == 1:
                    if cell.value is not None:
                        cell.value = str(cell.value).strip().upper()
                    cell.font = Font(name="Helvetica", size=12, bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if cell.value is not None:
                        max_widths[cell.column] = max(
                            max_widths[cell.column],
                            len(_display_value(cell.value, col_types.get(cell.column))),
                        )
                    continue
                col_type = col_types.get(cell.column)
                if not col_type:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if cell.value is not None:
                        max_widths[cell.column] = max(
                            max_widths[cell.column],
                            len(_display_value(cell.value, col_type)),
                        )
                    continue
                if _is_dash(cell.value):
                    cell.number_format = TEXT_FMT
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if cell.value is not None:
                        max_widths[cell.column] = max(
                            max_widths[cell.column],
                            len(_display_value(cell.value, col_type)),
                        )
                    continue
                if col_type == "date":
                    if isinstance(cell.value, str):
                        parsed = _parse_date(cell.value)
                        if parsed is not None:
                            cell.value = parsed
                    cell.number_format = DATE_FMT
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_type == "time":
                    if isinstance(cell.value, str):
                        parsed = _parse_time(cell.value)
                        if parsed is not None:
                            cell.value = parsed
                    cell.number_format = TIME_FMT
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_type == "currency":
                    if isinstance(cell.value, str):
                        parsed = _parse_float(cell.value)
                        if parsed is not None:
                            cell.value = parsed
                    cell.number_format = CURRENCY_FMT
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_type == "int":
                    if isinstance(cell.value, str):
                        parsed = _parse_float(cell.value)
                        if parsed is not None:
                            cell.value = int(parsed)
                    cell.number_format = INTEGER_FMT
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_type == "text":
                    cell.number_format = TEXT_FMT
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                if cell.value is not None:
                    max_widths[cell.column] = max(
                        max_widths[cell.column],
                        len(_display_value(cell.value, col_type)),
                    )

        for col_idx, width in max_widths.items():
            if width <= 0:
                continue
            # Add padding and scale a bit to better match font metrics.
            adjusted = max(int(width * 1.1) + 2, 8)
            # Keep a generous cap to avoid extreme outliers.
            adjusted = min(adjusted, 120)
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = adjusted
            ws.column_dimensions[col_letter].bestFit = True
            ws.column_dimensions[col_letter].auto_size = True
    wb.save(path)
